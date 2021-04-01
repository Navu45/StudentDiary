import openpyxl
from bs4 import BeautifulSoup
import re
import requests
import json
import datetime


def get_schedule_from_mirea():
    """Парсит расписаниие с 'https://www.mirea.ru/schedule/' и преобразует его в json файлы.
    На выходе получается 3 .xlsx файла и 3 .json файла (нумерация файлов в соответствии с курсом).
            """

    page = requests.get('https://www.mirea.ru/schedule/')
    soup = BeautifulSoup(page.text, "html.parser")

    result = soup.find('div', {'class': 'rasspisanie'}).find(string='Институт информационных технологий').find_parent(
        'div').find_parent('div').findAll('a', {'class': 'uk-link-toggle'})
    for x in result:
        if re.search('ИИТ.*xlsx', str(x)):
            if re.search('1к', str(x)):
                f = open('file1.xlsx', 'wb')
                filexlsx = requests.get(x['href'])
                f.write(filexlsx.content)
                f.close()
            if re.search('2к', str(x)):
                f = open('file2.xlsx', 'wb')
                filexlsx = requests.get(x['href'])
                f.write(filexlsx.content)
                f.close()
            if re.search('3к', str(x)):
                f = open('file3.xlsx', 'wb')
                filexlsx = requests.get(x['href'])
                f.write(filexlsx.content)
                f.close()

    for n in range(1, 4):
        book = openpyxl.load_workbook('file{}.xlsx'.format(n))
        sheet = book.active

        num_cols = sheet.max_column

        groups_list = []
        groups = {}
        week_days = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT']
        for col_index in range(1, num_cols + 1):
            group_cell = str(sheet.cell(2, col_index).value)
            if re.search(r'\w{4}-\d\d-\d\d', group_cell):
                groups_list.append(group_cell)
                week = {'MON': None, 'TUE': None, 'WED': None, 'THU': None, 'FRI': None, 'SAT': None}
                for k in range(6):
                    day = [[], [], [], [], [], []]
                    for i in range(6):
                        for j in range(2):
                            subject = sheet.cell(4 + j + i * 2 + k * 12, col_index).value
                            lesson_type = sheet.cell(4 + j + i * 2 + k * 12, col_index + 1).value
                            lecturer = sheet.cell(4 + j + i * 2 + k * 12, col_index + 2).value
                            classroom = sheet.cell(4 + j + i * 2 + k * 12, col_index + 3).value
                            url = sheet.cell(4 + j + i * 2 + k * 12, col_index + 4).value
                            lesson = {'subject': subject, 'lesson_type': lesson_type, 'lecturer': lecturer,
                                      'classroom': classroom, 'url': url}
                            day[i].append(lesson)
                    week[week_days[k]] = day
                groups.update({group_cell: week})

        with open("groups{}.json".format(n), "w") as write_file:
            json.dump(groups, write_file)


def show_schedule_for_day(group, d):
    """Парсит расписаниие json файл и возвращает расписание определенного дня в str.

            Параметры:
                    group (str): студенческая группа
                    d (datetime.datetime): дата

            Возвращаемое значение:
                    schedule (str): расписание на день
                """

    if d.weekday() == 6:
        return 'Занятий нет\n'
    course = -(int(group[-2::]) - int(
        str(datetime.datetime.now().year)[-2::])) if datetime.datetime.now().month < 7 else -(
            int(group[-2::]) - int(str(datetime.datetime.now().year)[-2::]) + 1)
    with open("groups{}.json".format(course), "r") as read_file:
        data = json.load(read_file)
    week_days = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT']
    schedule = ''
    for i in range(6):
        str1 = str(i + 1) + ') '
        if data[group][week_days[d.weekday()]][i][(d.isocalendar()[1] + 1) % 2]['subject']:
            str1 += str(data[group][week_days[d.weekday()]][i][(d.isocalendar()[1] + 1) % 2]['subject']) + ', '
            if data[group][week_days[d.weekday()]][i][(d.isocalendar()[1] + 1) % 2]['lesson_type']:
                str1 += str(data[group][week_days[d.weekday()]][i][(d.isocalendar()[1] + 1) % 2]['lesson_type']) + ', '
            else:
                str1 += '--, '
            if data[group][week_days[d.weekday()]][i][(d.isocalendar()[1] + 1) % 2]['lecturer']:
                str1 += str(data[group][week_days[d.weekday()]][i][(d.isocalendar()[1] + 1) % 2]['lecturer']) + ', '
            else:
                str1 += '--, '
            if data[group][week_days[d.weekday()]][i][(d.isocalendar()[1] + 1) % 2]['classroom'] != 'Д':
                if data[group][week_days[d.weekday()]][i][(d.isocalendar()[1] + 1) % 2]['classroom']:
                    str1 += str(
                        data[group][week_days[d.weekday()]][i][(d.isocalendar()[1] + 1) % 2]['classroom']) + '\n'
                else:
                    str1 += '--\n'
            elif data[group][week_days[d.weekday()]][i][(d.isocalendar()[1] + 1) % 2]['url']:
                str1 += data[group][week_days[d.weekday()]][i][(d.isocalendar()[1] + 1) % 2]['url'] + '\n'
            else:
                str1 += '--\n'
        else:
            str1 += '--\n'
        schedule += str1
    return schedule


def show_schedule_for_week(group, d):
    """Парсит расписаниие json файл и возвращает расписание на определенную неделю в str.

                Параметры:
                        group (str): студенческая группа
                        d (datetime.datetime): дата

                Возвращаемое значение:
                        schedule (str): расписание на неделю
                    """
    d1 = d - datetime.timedelta(days=d.weekday())
    schedule = ''
    week_days = ['понедельник', 'вторник', 'среду', 'четверг', 'пятница', 'субботу']
    months = ['января', 'февраля', 'марта', 'апреля', 'мая', 'июня', 'июля', 'августа', 'сентября', 'октября', 'ноября',
              'декабря']
    for i in range(6):
        schedule += 'Расписание на ' + week_days[i] + ' ' + str(d1.day) + ' ' + months[d1.month - 1] + ':\n'
        schedule += show_schedule_for_day(group, d1)
        d1 += datetime.timedelta(days=1)
    return schedule
